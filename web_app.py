"""
Billing Submission Automation - Web UI version.

Double-click this (once packaged as an .exe) and it starts a small local
web server, then opens the interface directly in your default browser
(Chrome). Upload the 3 files, click Process, download the result.

IMPORTANT: the actual report-building logic below (process_one and its
helpers) is copied UNCHANGED from the tested script - only the interface
around it changed, from a desktop Tkinter window to a local web page.

Requires: pip install flask pandas numpy openpyxl
Run directly:  python web_app.py
Build to exe:  see the instructions given alongside this file.
"""

import argparse
import contextlib
import re
import shutil
import socket
import sys
import tempfile
import threading
import traceback
import uuid
import webbrowser
import zipfile
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from flask import Flask, jsonify, request, send_file

# =======================================================================
# UNCHANGED PROCESSING LOGIC (identical to the tested script)
# =======================================================================

DATE_COLUMNS = {
    "DATEOFBIRTH", "DATEOFSERVICE", "DATE OF INJURY", "BILLDATE",
    "STATUS_CHANGED_ON", "BILLCREATEDDATE", "Bill Submission Date",
    "Payer Date", "response_date",
}
EXCEL_EPOCH = pd.Timestamp(1899, 12, 30)


def fast_read_excel(path: Path, sheet_name=0):
    try:
        return pd.read_excel(path, sheet_name=sheet_name, engine="calamine")
    except (ImportError, ValueError):
        pass

    if str(path).lower().endswith(".xlsb"):
        try:
            return pd.read_excel(path, sheet_name=sheet_name, engine="pyxlsb")
        except ImportError:
            raise ImportError(
                "This file is a .xlsb (Excel Binary Workbook). Reading it needs "
                "either `pip install python-calamine` or `pip install pyxlsb`."
            )

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    data = list(rows_iter)
    wb.close()
    return pd.DataFrame(data, columns=header)


def col_letter(n: int) -> str:
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def letter_to_num(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def xml_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def fmt_num(v: float) -> str:
    v = float(v)
    if v == int(v):
        return str(int(v))
    return repr(v)


def process_one(bdr_path: Path, case_ar_path: Path, master_path: Path, output_path: Path):
    print(f"\n=== Processing ===\n  BDR:     {bdr_path}\n  Case_AR: {case_ar_path}\n  Master:  {master_path}\n  Output:  {output_path}")
    t0 = datetime.now()

    work_dir = Path(tempfile.gettempdir()) / ("_xlsx_work_" + output_path.stem)
    if work_dir.exists():
        shutil.rmtree(work_dir)
    work_dir.mkdir(parents=True)

    with zipfile.ZipFile(master_path) as z:
        z.extractall(work_dir)

    sheet1_path = work_dir / "xl/worksheets/sheet1.xml"
    sst_path = work_dir / "xl/sharedStrings.xml"
    workbook_path = work_dir / "xl/workbook.xml"
    ct_path = work_dir / "[Content_Types].xml"
    rels_path = work_dir / "xl/_rels/workbook.xml.rels"
    calc_chain_path = work_dir / "xl/calcChain.xml"

    sheet1_xml = sheet1_path.read_text(encoding="utf-8")

    header_row_m = re.search(r'<row r="1"[^>]*>(.*?)</row>', sheet1_xml, re.S)
    header_cells = re.findall(r'<c r="([A-Z]+)1"[^>]*t="s"[^>]*><v>(\d+)</v></c>', header_row_m.group(1))

    sst_xml = sst_path.read_text(encoding="utf-8")
    sst_items = re.findall(r"<si>(.*?)</si>", sst_xml, re.S)

    def sst_plain_text(si_inner: str) -> str:
        return "".join(re.findall(r"<t[^>]*>(.*?)</t>", si_inner, re.S))

    master_headers = {}
    for letter, idx in header_cells:
        master_headers[letter] = sst_plain_text(sst_items[int(idx)])

    master_cols_ordered = sorted(master_headers.items(), key=lambda kv: letter_to_num(kv[0]))
    print(f"  Master header columns: {len(master_cols_ordered)}")

    cols_block = re.search(r"<cols>(.*?)</cols>", sheet1_xml, re.S).group(1)
    col_style = {}
    for mn, mx, st in re.findall(r'<col min="(\d+)" max="(\d+)"[^>]*style="(\d+)"', cols_block):
        for c in range(int(mn), int(mx) + 1):
            col_style[c] = st

    bdr = pd.read_csv(bdr_path, dtype=str, keep_default_na=True, low_memory=False)
    bdr.columns = [c.strip() for c in bdr.columns]
    n = len(bdr)
    print(f"  BDR rows: {n}, cols: {len(bdr.columns)}")

    # Date-column matching is case-insensitive (some providers spell these
    # differently) and parsed_dates is keyed by the ACTUAL BDR column name.
    date_columns_norm = {d.strip().lower() for d in DATE_COLUMNS}
    parsed_dates = {}
    for actual_col in bdr.columns:
        if actual_col.strip().lower() in date_columns_norm:
            parsed_dates[actual_col] = pd.to_datetime(bdr[actual_col], errors="coerce", format="mixed")

    bdr_col_by_norm = {c.strip().lower(): c for c in bdr.columns}
    bill_sub_date_col = bdr_col_by_norm.get("bill submission date")
    if bill_sub_date_col is None:
        raise ValueError("Could not find a 'Bill Submission Date' column in the BDR file.")

    case_ar = fast_read_excel(case_ar_path, sheet_name=0)
    case_ar["CASEID"] = case_ar["CASEID"].astype(str).str.strip()
    edi_lookup = dict(zip(case_ar["CASEID"], case_ar["EDI STATUS"]))
    print(f"  [{ (datetime.now()-t0).total_seconds():.1f}s] Case_AR loaded: {len(case_ar)} rows")

    bdr_caseid = bdr["CASEID"].astype(str).str.strip()

    status_norm = bdr["STATUS"].astype(str).str.strip().str.lower()
    has_sub_date = ~parsed_dates[bill_sub_date_col].isna()
    raw_verif_norm = bdr_caseid.map(edi_lookup)
    raw_verif_norm = raw_verif_norm.apply(
        lambda v: str(v).strip().lower() if v is not None and not (isinstance(v, float) and pd.isna(v)) else ""
    )
    insurance_present = bdr["INSURANCE NAME"].notna() & (bdr["INSURANCE NAME"].str.strip() != "")

    CASCADE = {
        "close": ("Close", "Bill", "Closed", "Closed"),
        "__has_sub_date__": ("Verified", "Bill", "Electronic", "Processed"),
        "verified": ("Verified", "Bill", "Electronic", "Ready to Bill"),
        "missing insurance": ("Missing Insurance", "Problem Case", "Missing Insurance", "Problem Case"),
        "waiting for verification": ("Waiting for Verification", "Problem Case", "Unverified", "Problem Case"),
        "__insurance_fallback__": ("Waiting for Verification", "Problem Case", "Unverified", "Problem Case"),
        "__blank__": (None, None, None, None),
    }

    def classify(i):
        if status_norm.iat[i] == "close":
            return CASCADE["close"]
        
        if has_sub_date.iat[i]:
            return CASCADE["__has_sub_date__"]

        rv = raw_verif_norm.iat[i]

        if rv in ("verified", "missing insurance", "waiting for verification"):
            return CASCADE[rv]

    # NEW CONDITION:
    # Submission Date is blank
    # Verification Status is blank
    # Insurance Name is blank
        if not has_sub_date.iat[i] and not rv and not insurance_present.iat[i]:
            return CASCADE["missing insurance"]

    # Insurance Name exists but Verification Status is blank
        if insurance_present.iat[i]:
            return CASCADE["__insurance_fallback__"]
        
        return CASCADE["__blank__"]

    classified = [classify(i) for i in range(n)]
    verif_list = [c[0] for c in classified]
    billtype_list = [c[1] for c in classified]
    editype_list = [c[2] for c in classified]
    edistatus_list = [c[3] for c in classified]
    print(f"  [{ (datetime.now()-t0).total_seconds():.1f}s] classification done")

    service_type = bdr["Service Type"]
    for i in range(n):
        if edistatus_list[i] == "Processed":
            sv = service_type.iat[i]
            if not (pd.isna(sv) or str(sv).strip() == ""):
                editype_list[i] = str(sv).strip()

    appt_series = bdr["APPOINTMENTTYPE"]
    suffix_pattern = re.compile(r"(\s+[xX]\d+|\s*-\d+)\s*$")

    def appt_base_key(v):
        if pd.isna(v):
            return None
        base = suffix_pattern.sub("", v).strip()
        base = re.sub(r"\(OC\)\s*$", "(O&O-Chiro)", base)
        return base

    appt_bases = appt_series.map(appt_base_key)
    appt_canonical_by_base = {}
    for base, val in zip(appt_bases, appt_series):
        if base is not None and base not in appt_canonical_by_base:
            appt_canonical_by_base[base] = val
    unique_appt_list = [appt_canonical_by_base[b] if b is not None else None for b in appt_bases]

    classified_map = {
        "Verification Status": verif_list,
        "Bill Type": billtype_list,
        "EDI service Type": editype_list,
        "EDI Status": edistatus_list,
        "UNIQUE APPOINTMENTTYPE": unique_appt_list,
    }

    uniq_count_m = re.search(r'uniqueCount="(\d+)"', sst_xml)
    next_sst_index = int(uniq_count_m.group(1))
    new_sst_entries = []
    sst_new_lookup = {}

    def get_string_index(s: str) -> int:
        nonlocal next_sst_index
        if s in sst_new_lookup:
            return sst_new_lookup[s]
        idx = next_sst_index
        sst_new_lookup[s] = idx
        new_sst_entries.append(s)
        next_sst_index += 1
        return idx

    # ---- Name matching is CASE-INSENSITIVE + trimmed, because different
    #      provider Master templates spell these headers differently
    #      (e.g. "EDI Service Type" vs "EDI service Type"). --------------
    def norm(s: str) -> str:
        return s.strip().lower()

    classified_map_norm = {norm(k): v for k, v in classified_map.items()}
    bdr_col_by_norm = {norm(c): c for c in bdr.columns}

    # ---- Which column actually holds BILLCREATEDDATE in THIS template?
    #      (was hardcoded to "BE" before - wrong whenever a template has a
    #      different column count/order, e.g. no "Market_Values" column) --
    billcreated_letter = None
    for letter, name in master_cols_ordered:
        if norm(name) == "billcreateddate":
            billcreated_letter = letter
            break
    if billcreated_letter is None:
        raise ValueError(
            "Could not find a 'BILLCREATEDDATE' column in this Master template - "
            "Billing Month/Billing Week formulas need it to know which cell to reference."
        )
    print(f"  BILLCREATEDDATE column detected as: {billcreated_letter}")

    bdr_col_set = set(bdr.columns)
    plan = []
    for letter, name in master_cols_ordered:
        name_norm = norm(name)
        if name_norm == "billing month":
            plan.append((letter, name, "formula_month", None))
        elif name_norm == "billing week":
            plan.append((letter, name, "formula_week", None))
        elif name_norm in classified_map_norm:
            plan.append((letter, name, "classified", classified_map_norm[name_norm]))
        elif name_norm == "submission date":
            plan.append((letter, name, "submission_date", None))
        elif name_norm in bdr_col_by_norm:
            plan.append((letter, name, "bdr", bdr_col_by_norm[name_norm]))
        else:
            plan.append((letter, name, None, None))

    plan_letters = [p[0] for p in plan]
    plan_kinds = [p[2] for p in plan]
    plan_styles = [col_style.get(letter_to_num(p[0]), "0") for p in plan]
    n_cols = len(plan)
    CHUNK_SIZE = 50_000

    def build_chunk_rows(start: int, end: int):
        local_n = end - start
        col_payloads_chunk = []
        for letter, name, kind, source in plan:
            if kind == "bdr":
                raw = bdr[source].iloc[start:end]
                if source in parsed_dates:
                    dt = parsed_dates[source].iloc[start:end]
                    serial = (dt - EXCEL_EPOCH) / pd.Timedelta(days=1)
                    payload = [None] * local_n
                    valid_idx = np.flatnonzero(~serial.isna().values)
                    sv = serial.values
                    for i in valid_idx:
                        payload[i] = ("n", fmt_num(sv[i]))
                else:
                    stripped = raw.str.strip()
                    notna_mask = raw.notna() & (stripped != "")
                    numeric = pd.to_numeric(raw, errors="coerce")
                    is_numeric = numeric.notna()
                    is_text = notna_mask & ~is_numeric

                    payload = [None] * local_n
                    num_idx = np.flatnonzero(is_numeric.values)
                    nv = numeric.values
                    for i in num_idx:
                        payload[i] = ("n", fmt_num(nv[i]))

                    text_idx = np.flatnonzero(is_text.values)
                    if len(text_idx) > 0:
                        text_vals = stripped.values
                        uniques = pd.unique(text_vals[text_idx])
                        idx_map = {v: get_string_index(v) for v in uniques}
                        for i in text_idx:
                            payload[i] = ("s", idx_map[text_vals[i]])
                col_payloads_chunk.append(payload)

            elif kind == "classified":
                vals = source[start:end]
                payload = [None] * local_n
                uniq_vals = set(v for v in vals if v)
                idx_map = {v: get_string_index(v) for v in uniq_vals}
                for i, v in enumerate(vals):
                    if v:
                        payload[i] = ("s", idx_map[v])
                col_payloads_chunk.append(payload)

            elif kind == "submission_date":
                dt = parsed_dates[bill_sub_date_col].iloc[start:end]
                serial = (dt - EXCEL_EPOCH) / pd.Timedelta(days=1)
                payload = [None] * local_n
                valid_idx = np.flatnonzero(~serial.isna().values)
                sv = serial.values
                for i in valid_idx:
                    payload[i] = ("n", fmt_num(sv[i]))
                col_payloads_chunk.append(payload)

            else:
                col_payloads_chunk.append(None)

        rows_out = []
        for i in range(local_n):
            r = start + i + 2
            parts = []
            for j in range(n_cols):
                kind = plan_kinds[j]
                letter = plan_letters[j]
                style = plan_styles[j]
                if kind == "formula_month":
                    parts.append(f'<c r="{letter}{r}" s="{style}"><f>IF({billcreated_letter}{r}="","",TEXT({billcreated_letter}{r},"MMM-YY"))</f></c>')
                elif kind == "formula_week":
                    parts.append(f'<c r="{letter}{r}" s="{style}"><f>IF({billcreated_letter}{r}="","",CONCATENATE("Week 0",INT((DAY({billcreated_letter}{r})-1)/7)+1))</f></c>')
                elif kind is None:
                    continue
                else:
                    cell = col_payloads_chunk[j][i]
                    if cell is None:
                        continue
                    t, val = cell
                    if t == "n":
                        parts.append(f'<c r="{letter}{r}" s="{style}"><v>{val}</v></c>')
                    else:
                        parts.append(f'<c r="{letter}{r}" s="{style}" t="s"><v>{val}</v></c>')
            rows_out.append(f'<row r="{r}" spans="1:{n_cols}">' + "".join(parts) + "</row>")
        return rows_out

    header_row_full = re.search(r'<row r="1"[^>]*>.*?</row>', sheet1_xml, re.S).group(0)
    m_sd = re.search(r"<sheetData>.*?</sheetData>", sheet1_xml, re.S)
    before_sd = sheet1_xml[: m_sd.start()]
    after_sd = sheet1_xml[m_sd.end():]

    last_row = n + 1
    last_col_letter = master_cols_ordered[-1][0]
    before_sd = re.sub(
        r'<dimension ref="[^"]*"/>',
        f'<dimension ref="A1:{last_col_letter}{last_row}"/>',
        before_sd,
    )

    with open(sheet1_path, "w", encoding="utf-8") as f:
        f.write(before_sd)
        f.write("<sheetData>")
        f.write(header_row_full)
        for start in range(0, n, CHUNK_SIZE):
            end = min(start + CHUNK_SIZE, n)
            rows_out = build_chunk_rows(start, end)
            f.write("".join(rows_out))
            print(f"  [{(datetime.now()-t0).total_seconds():.1f}s] rows {start}-{end} written")
        f.write("</sheetData>")
        f.write(after_sd)

    wb_xml = workbook_path.read_text(encoding="utf-8")
    wb_xml = re.sub(
        r'(_xlnm\._FilterDatabase" localSheetId="0" hidden="1">Data!\$A\$1:\$)[A-Z]+\$\d+',
        rf"\g<1>{last_col_letter}${last_row}",
        wb_xml,
    )
    if "fullCalcOnLoad" not in wb_xml:
        wb_xml = wb_xml.replace("<calcPr calcId=", '<calcPr fullCalcOnLoad="1" calcId=')
    workbook_path.write_text(wb_xml, encoding="utf-8")

    if calc_chain_path.exists():
        calc_chain_path.unlink()
        ct_xml = ct_path.read_text(encoding="utf-8")
        ct_xml = ct_xml.replace(
            '<Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/>',
            "",
        )
        ct_path.write_text(ct_xml, encoding="utf-8")
        rels_xml = rels_path.read_text(encoding="utf-8")
        rels_xml = re.sub(r'<Relationship Id="rId10"[^/]*calcChain[^/]*/>', "", rels_xml)
        rels_path.write_text(rels_xml, encoding="utf-8")

    if new_sst_entries:
        additions = "".join(f'<si><t xml:space="preserve">{xml_escape(s)}</t></si>' for s in new_sst_entries)
        sst_xml_new = sst_xml.replace("</sst>", additions + "</sst>")
        total_count = int(re.search(r'count="(\d+)"', sst_xml).group(1)) + len(new_sst_entries)
        sst_xml_new = re.sub(
            r'count="\d+" uniqueCount="\d+"',
            f'count="{total_count}" uniqueCount="{next_sst_index}"',
            sst_xml_new,
            count=1,
        )
        sst_path.write_text(sst_xml_new, encoding="utf-8")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        output_path.unlink()
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in work_dir.rglob("*"):
            if f.is_file():
                zf.write(f, f.relative_to(work_dir))

    shutil.rmtree(work_dir)
    dt_elapsed = (datetime.now() - t0).total_seconds()
    print(f"  Done: {n} rows -> {output_path}  ({dt_elapsed:.1f}s)")


# =======================================================================
# WEB SERVER / UI (this is the new part - opens in Chrome, not VS Code)
# =======================================================================

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024 * 1024  # 2 GB upload ceiling

JOBS = {}  # job_id -> {"status": "running"|"done"|"error", "log": [str], "output_path": str|None}


class _ListWriter:
    def __init__(self, buf):
        self.buf = buf

    def write(self, text):
        if text:
            self.buf.append(text)

    def flush(self):
        pass


def _run_job(job_id, bdr_path, case_ar_path, master_path, output_path):
    job = JOBS[job_id]
    writer = _ListWriter(job["log"])
    try:
        with contextlib.redirect_stdout(writer), contextlib.redirect_stderr(writer):
            process_one(bdr_path, case_ar_path, master_path, output_path)
        job["status"] = "done"
        job["output_path"] = str(output_path)
    except Exception:
        job["log"].append(traceback.format_exc())
        job["status"] = "error"


@app.route("/")
def index():
    return INDEX_HTML


@app.route("/api/process", methods=["POST"])
def api_process():
    bdr = request.files.get("bdr")
    case_ar = request.files.get("case_ar")
    master = request.files.get("master")
    if not (bdr and case_ar and master):
        return jsonify({"error": "All three files are required."}), 400

    job_id = uuid.uuid4().hex
    work = Path(tempfile.mkdtemp(prefix="billing_upload_"))
    bdr_path = work / bdr.filename
    case_ar_path = work / case_ar.filename
    master_path = work / master.filename
    bdr.save(bdr_path)
    case_ar.save(case_ar_path)
    master.save(master_path)
    output_path = work / "Billing_Submission_Master_File.xlsx"

    JOBS[job_id] = {"status": "running", "log": [], "output_path": None}
    t = threading.Thread(
        target=_run_job,
        args=(job_id, bdr_path, case_ar_path, master_path, output_path),
        daemon=True,
    )
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def api_status(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "unknown job"}), 404
    return jsonify({
        "status": job["status"],
        "log": "".join(job["log"]),
        "ready": job["status"] == "done",
    })


@app.route("/api/download/<job_id>")
def api_download(job_id):
    job = JOBS.get(job_id)
    if not job or job["status"] != "done":
        return "Not ready", 404
    return send_file(job["output_path"], as_attachment=True, download_name="Billing_Submission_Master_File.xlsx")


def _find_free_port() -> int:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(("127.0.0.1", 0))
    port = s.getsockname()[1]
    s.close()
    return port


def main():
    port = _find_free_port()

    def open_browser():
        import time
        time.sleep(0.8)
        webbrowser.open(f"http://127.0.0.1:{port}")

    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False)


# =======================================================================
# EMBEDDED FRONTEND (single file - no external assets, works fully offline)
# =======================================================================

INDEX_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Billing Submission Automation</title>
<style>
  :root{
    --bg:#0A0E16;
    --panel:#121826;
    --panel-border:#212B3D;
    --text:#EAEEF6;
    --text-dim:#8B93A7;
    --accent-a:#4F6EF7;
    --accent-b:#8B5CF6;
    --teal:#2DD4BF;
    --danger:#F87171;
    --mono: "Cascadia Code","Consolas","SFMono-Regular",Menlo,monospace;
    --sans: -apple-system,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
  }
  *{box-sizing:border-box;}
  body{
    margin:0; min-height:100vh; background:
      radial-gradient(1200px 600px at 15% -10%, rgba(79,110,247,.14), transparent 60%),
      radial-gradient(1000px 500px at 100% 0%, rgba(45,212,191,.10), transparent 55%),
      var(--bg);
    color:var(--text); font-family:var(--sans);
    padding:28px 20px 60px;
  }
  .wrap{max-width:1040px; margin:0 auto;}
  .hero{display:flex; align-items:center; justify-content:space-between; margin-bottom:26px;}
  .hero-left{display:flex; align-items:center; gap:14px;}
  .logo{
    width:44px; height:44px; border-radius:12px; flex:0 0 auto;
    background:linear-gradient(135deg,var(--accent-a),var(--accent-b));
    display:flex; align-items:center; justify-content:center;
    box-shadow:0 8px 24px -8px rgba(79,110,247,.6);
  }
  .logo svg{width:22px; height:22px;}
  h1{font-size:21px; margin:0; font-weight:650; letter-spacing:-.01em;}
  .sub{color:var(--text-dim); font-size:13px; margin-top:3px;}
  .badge{
    display:flex; align-items:center; gap:7px;
    border:1px solid var(--panel-border); background:rgba(45,212,191,.07);
    color:var(--teal); font-size:12.5px; padding:7px 13px; border-radius:999px;
    white-space:nowrap;
  }
  .badge .dot{width:6px; height:6px; border-radius:50%; background:var(--teal); box-shadow:0 0 8px var(--teal);}

  .grid{display:grid; grid-template-columns:1.35fr 1fr; gap:18px;}
  @media (max-width:860px){ .grid{grid-template-columns:1fr;} }

  .card{
    background:var(--panel); border:1px solid var(--panel-border); border-radius:16px;
    padding:22px;
  }
  .card h2{font-size:14.5px; margin:0 0 3px; font-weight:650;}
  .card .hint{color:var(--text-dim); font-size:12.5px; margin:0 0 16px;}

  .file-row{
    display:flex; align-items:center; gap:12px;
    border:1px solid var(--panel-border); border-radius:12px; padding:12px 14px; margin-bottom:10px;
    background:rgba(255,255,255,.015);
  }
  .file-icon{
    width:34px; height:34px; border-radius:9px; flex:0 0 auto;
    background:rgba(79,110,247,.12); display:flex; align-items:center; justify-content:center;
  }
  .file-icon svg{width:16px; height:16px; opacity:.9;}
  .file-info{flex:1; min-width:0;}
  .file-name{font-size:13.5px; font-weight:560; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;}
  .file-meta{font-size:11.5px; color:var(--text-dim); margin-top:1px;}
  .file-meta.ok{color:var(--teal);}
  .file-meta.req{color:#F0B429;}
  .upload-btn{
    border:1px solid var(--panel-border); background:rgba(255,255,255,.03); color:var(--text);
    font-size:12.5px; font-weight:600; padding:8px 14px; border-radius:9px; cursor:pointer;
    white-space:nowrap;
  }
  .upload-btn:hover{background:rgba(255,255,255,.07);}

  .process-btn{
    width:100%; margin-top:6px; padding:14px; border:none; border-radius:12px;
    font-size:14px; font-weight:650; letter-spacing:.02em; color:#fff; cursor:pointer;
    background:linear-gradient(135deg,var(--accent-a),var(--accent-b));
    box-shadow:0 10px 26px -10px rgba(124,92,252,.55);
    transition:transform .12s ease, box-shadow .12s ease, opacity .12s ease;
  }
  .process-btn:hover:not(:disabled){transform:translateY(-1px);}
  .process-btn:disabled{opacity:.5; cursor:not-allowed; box-shadow:none;}

  .out-row{
    display:flex; align-items:center; gap:12px;
    border:1px solid var(--panel-border); border-radius:12px; padding:12px 14px; margin-bottom:16px;
    background:rgba(255,255,255,.015);
  }
  .out-name{font-size:13px; flex:1; min-width:0; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; color:var(--text-dim);}
  .download-btn{
    border:1px solid var(--panel-border); background:rgba(45,212,191,.10); color:var(--teal);
    font-size:12.5px; font-weight:650; padding:8px 14px; border-radius:9px; cursor:pointer;
  }
  .download-btn:disabled{opacity:.35; cursor:not-allowed;}

  .status-line{display:flex; align-items:center; gap:8px; font-size:13px; margin-bottom:14px;}
  .status-dot{width:8px; height:8px; border-radius:50%; background:var(--text-dim);}
  .status-dot.running{background:#F0B429; box-shadow:0 0 8px #F0B429; animation:pulse 1.1s infinite;}
  .status-dot.done{background:var(--teal); box-shadow:0 0 8px var(--teal);}
  .status-dot.error{background:var(--danger); box-shadow:0 0 8px var(--danger);}
  @keyframes pulse{0%,100%{opacity:1;}50%{opacity:.35;}}

  .clear-btn{
    border:1px solid var(--panel-border); background:transparent; color:var(--text-dim);
    font-size:12.5px; padding:8px 14px; border-radius:9px; cursor:pointer;
  }
  .clear-btn:hover{color:var(--text); background:rgba(255,255,255,.04);}

  .log-card{margin-top:18px;}
  .log-box{
    background:#080B12; border:1px solid var(--panel-border); border-radius:12px;
    padding:14px 16px; height:230px; overflow-y:auto;
    font-family:var(--mono); font-size:12.5px; line-height:1.6; color:#B9C2D4;
    white-space:pre-wrap; word-break:break-word;
  }
  .footnote{color:var(--text-dim); font-size:12px; margin-top:12px;}
  input[type=file]{display:none;}
</style>
</head>
<body>
<div class="wrap">

  <div class="hero">
    <div class="hero-left">
      <div class="logo">
        <svg viewBox="0 0 24 24" fill="none"><path d="M13 2 4 14h6l-1 8 9-12h-6l1-8Z" fill="white"/></svg>
      </div>
      <div>
        <h1>Billing Submission Automation</h1>
        <div class="sub">BDR + Case_AR + Master &rarr; Billing Submission Master File</div>
      </div>
    </div>
    <div class="badge"><span class="dot"></span>Runs on your machine &mdash; files never leave this PC</div>
  </div>

  <div class="grid">
    <div class="card">
      <h2>Input Files</h2>
      <div class="hint">Provider and filename do not matter &mdash; just pick the right file for each row.</div>

      <div class="file-row" id="row-bdr">
        <div class="file-icon"><svg viewBox="0 0 24 24" fill="none" stroke="#4F6EF7" stroke-width="1.6"><path d="M6 2h9l5 5v15H6z"/><path d="M14 2v6h6"/></svg></div>
        <div class="file-info">
          <div class="file-name" id="name-bdr">No BDR file selected</div>
          <div class="file-meta req" id="meta-bdr">Required &middot; .csv</div>
        </div>
        <button class="upload-btn" onclick="document.getElementById('input-bdr').click()">Upload</button>
        <input type="file" id="input-bdr" accept=".csv">
      </div>

      <div class="file-row" id="row-case_ar">
        <div class="file-icon"><svg viewBox="0 0 24 24" fill="none" stroke="#4F6EF7" stroke-width="1.6"><path d="M6 2h9l5 5v15H6z"/><path d="M14 2v6h6"/></svg></div>
        <div class="file-info">
          <div class="file-name" id="name-case_ar">No Case_AR file selected</div>
          <div class="file-meta req" id="meta-case_ar">Required &middot; .xlsx</div>
        </div>
        <button class="upload-btn" onclick="document.getElementById('input-case_ar').click()">Upload</button>
        <input type="file" id="input-case_ar" accept=".xlsx,.xlsm">
      </div>

      <div class="file-row" id="row-master">
        <div class="file-icon"><svg viewBox="0 0 24 24" fill="none" stroke="#4F6EF7" stroke-width="1.6"><path d="M6 2h9l5 5v15H6z"/><path d="M14 2v6h6"/></svg></div>
        <div class="file-info">
          <div class="file-name" id="name-master">No Master file selected</div>
          <div class="file-meta req" id="meta-master">Required &middot; .xlsx</div>
        </div>
        <button class="upload-btn" onclick="document.getElementById('input-master').click()">Upload</button>
        <input type="file" id="input-master" accept=".xlsx,.xlsm">
      </div>

      <button class="process-btn" id="process-btn" disabled onclick="startProcessing()">PROCESS FILES</button>
    </div>

    <div class="card">
      <h2>Output</h2>
      <div class="hint">Nothing is downloaded until processing finishes.</div>

      <div class="out-row">
        <div class="out-name" id="out-name">Billing_Submission_Master_File.xlsx</div>
        <button class="download-btn" id="download-btn" disabled onclick="downloadResult()">Download</button>
      </div>

      <div class="status-line">
        <span class="status-dot" id="status-dot"></span>
        <span id="status-text">Ready</span>
      </div>

      <button class="clear-btn" onclick="clearAll()">Clear</button>

      <div class="footnote">After downloading, open the file in Excel and use Data &rarr; Refresh All so the Summary sheet picks up the new data.</div>
    </div>
  </div>

  <div class="card log-card">
    <h2>Processing Log</h2>
    <div class="hint">Live output from the report engine.</div>
    <div class="log-box" id="log-box">Ready. Select the three input files.</div>
  </div>

</div>

<script>
const files = { bdr: null, case_ar: null, master: null };
let currentJob = null;
let pollTimer = null;

for (const key of ['bdr','case_ar','master']) {
  document.getElementById('input-' + key).addEventListener('change', (e) => {
    const f = e.target.files[0];
    if (!f) return;
    files[key] = f;
    document.getElementById('name-' + key).textContent = f.name;
    const meta = document.getElementById('meta-' + key);
    meta.textContent = 'Selected \u00b7 ' + (f.size/1024/1024).toFixed(2) + ' MB';
    meta.className = 'file-meta ok';
    updateProcessButton();
  });
}

function updateProcessButton(){
  const ready = files.bdr && files.case_ar && files.master;
  document.getElementById('process-btn').disabled = !ready;
}

function setStatus(kind, text){
  const dot = document.getElementById('status-dot');
  dot.className = 'status-dot' + (kind ? ' ' + kind : '');
  document.getElementById('status-text').textContent = text;
}

function appendLog(text){
  const box = document.getElementById('log-box');
  box.textContent = text;
  box.scrollTop = box.scrollHeight;
}

async function startProcessing(){
  document.getElementById('process-btn').disabled = true;
  document.getElementById('download-btn').disabled = true;
  setStatus('running', 'Processing...');
  appendLog('Uploading files...');

  const form = new FormData();
  form.append('bdr', files.bdr);
  form.append('case_ar', files.case_ar);
  form.append('master', files.master);

  try {
    const res = await fetch('/api/process', { method: 'POST', body: form });
    const data = await res.json();
    if (!res.ok) throw new Error(data.error || 'Upload failed');
    currentJob = data.job_id;
    pollTimer = setInterval(pollStatus, 700);
  } catch (err) {
    setStatus('error', 'Failed');
    appendLog('Error: ' + err.message);
    document.getElementById('process-btn').disabled = false;
  }
}

async function pollStatus(){
  if (!currentJob) return;
  const res = await fetch('/api/status/' + currentJob);
  const data = await res.json();
  appendLog(data.log || 'Processing...');
  if (data.status === 'done') {
    clearInterval(pollTimer);
    setStatus('done', 'Completed');
    document.getElementById('download-btn').disabled = false;
    document.getElementById('process-btn').disabled = false;
  } else if (data.status === 'error') {
    clearInterval(pollTimer);
    setStatus('error', 'Failed');
    document.getElementById('process-btn').disabled = false;
  }
}

function downloadResult(){
  if (!currentJob) return;
  window.location.href = '/api/download/' + currentJob;
}

function clearAll(){
  if (pollTimer) clearInterval(pollTimer);
  currentJob = null;
  files.bdr = files.case_ar = files.master = null;
  for (const key of ['bdr','case_ar','master']) {
    document.getElementById('name-' + key).textContent = 'No ' + (key==='case_ar'?'Case_AR':key[0].toUpperCase()+key.slice(1)) + ' file selected';
    const meta = document.getElementById('meta-' + key);
    meta.textContent = 'Required \u00b7 ' + (key==='bdr' ? '.csv' : '.xlsx');
    meta.className = 'file-meta req';
    document.getElementById('input-' + key).value = '';
  }
  updateProcessButton();
  document.getElementById('download-btn').disabled = true;
  setStatus('', 'Ready');
  appendLog('Ready. Select the three input files.');
}
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()